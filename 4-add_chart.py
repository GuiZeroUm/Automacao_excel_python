from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1- Lê pasta de trabalho e planilha
wb = load_workbook("./data/pivot_table.xlsx")
sheet = wb["Relatorio"]

# 2- Referencias das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3- Adicionando dados e categorias no grafico
barchart = BarChart()

data = Reference(
    sheet,
    min_col = min_column+1,
    max_col = max_column,
    min_row = min_row,
    max_row = max_row
)

categories = Reference(
    sheet,
    min_col = min_column,
    max_col = min_column,
    min_row = min_row+1,
    max_row = max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# 4- criando o grafico
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2

# 5- salvando o workbook
wb.save("./data/barchart.xlsx")